import sys
import numpy as np
import os
import cv2
import itertools
import glob
import random
from tqdm import tqdm
import tensorflow as tf
from datetime import datetime
from scipy.io import loadmat
from openpyxl import Workbook
from openpyxl.styles import PatternFill


class MessageType:
    HEADER = '\033[95m'
    OKBLUE = '\033[94m'
    OKGREEN = '\033[92m'
    WARNING = '\033[93m'
    FAIL = '\033[91m'
    ENDC = '\033[0m'
    BOLD = '\033[1m'
    UNDERLINE = '\033[4m'


class FilesName:
    TRAIN                                   = "[TRAIN]                                      "
    EVAL                                    = "[EVAL]                                       "
    VIS                                     = "[VIS]                                        "
    GENERAL_UTILS                           = "[GENERAL_UTILS]                              "


def print_message(filename, message, mess_type=""):
    if mess_type == "":
        print(filename, MessageType.ENDC, message, MessageType.ENDC)
    else:
        print(filename, mess_type, message, MessageType.ENDC)


def map_parts_to_classes(num_classes, num_parts):
    """
    :param num_classes: number of classes
    :param num_parts: map the num_parts parts into the num_classes classes
    :return: a map from 21 classes to 'num_parts' parts
    """
    map_pc = []

    if num_classes == 2:
        if num_parts == 2:
        # Background-Person
            map_pc = [
                [0, 1],  # Background
                [1, 2],  # Person
            ]

        elif num_parts == 7:
            # Person parts
            map_pc = [
                [0, 1],  # Background
                [1, 7],  # Person
            ]

        else:
            print_message(FilesName.GENERAL_UTILS, "Error: num_classes = " + str(num_classes) +
                          " and num_parts = " + str(num_parts), MessageType.FAIL)
            exit(-1)

    elif num_classes == 21:

        if num_parts == 21:
            # 21 classes
            map_pc = [
                [0, 1],  # Background
                [1, 2],  # Aeroplane
                [2, 3],  # Bycicle
                [3, 4],  # Bird
                [4, 5],  # Boat
                [5, 6],  # Bottle
                [6, 7],  # Bus
                [7, 8],  # Car
                [8, 9],  # Cat
                [9, 10],  # Chair
                [10, 11],  # Cow
                [11, 12],  # Dining table
                [12, 13],  # Dog
                [13, 14],  # Horse
                [14, 15],  # Motorbike
                [15, 16],  # Person
                [16, 17],  # Potted plant
                [17, 18],  # Sheep
                [18, 19],  # Sofa
                [19, 20],  # Train
                [20, 21],  # Tv monitor
            ]

        elif num_parts == 58:
            # 58 parts
            map_pc = [
                [0, 1],  # Background
                [1, 6],  # Aeroplane
                [6, 8],  # Bycicle
                [8, 12],  # Bird
                [12, 13],  # Boat
                [13, 15],  # Bottle
                [15, 18],  # Bus
                [18, 23],  # Car
                [23, 27],  # Cat
                [27, 28],  # Chair
                [28, 32],  # Cow
                [32, 33],  # Dining table
                [33, 37],  # Dog
                [37, 41],  # Horse
                [41, 43],  # Motorbike
                [43, 49],  # Person
                [49, 51],  # Potted plant
                [51, 54],  # Sheep
                [54, 55],  # Sofa
                [55, 56],  # Train
                [56, 58],  # Tv monitor
            ]

        elif num_parts == 108:
            # 108 parts
            map_pc = [
                [0, 1],  # Background
                [1, 6],  # Aeroplane
                [6, 10],  # Bycicle
                [10, 18],  # Bird
                [18, 19],  # Boat
                [19, 21],  # Bottle
                [21, 29],  # Bus
                [29, 36],  # Car
                [36, 45],  # Cat
                [45, 46],  # Chair
                [46, 54],  # Cow
                [54, 55],  # Dining table
                [55, 65],  # Dog
                [65, 73],  # Horse
                [73, 77],  # Motorbike
                [77, 89],  # Person
                [89, 91],  # Potted plant
                [91, 99],  # Sheep
                [99, 100],  # Sofa
                [100, 107],  # Train
                [107, 108],  # Tv monitor
            ]

    else:
        print_message(FilesName.GENERAL_UTILS, "Error: num_classes = " + str(num_classes) +
                      " and num_parts = " + str(num_parts), MessageType.FAIL)
        exit(-1)

    return map_pc


def list_parts_names(num_parts):
    """
    :return: list parts name
    """

    list_parts = []
    if num_parts == 108:
        list_parts = ['background', 'aeroplane_body', 'aeroplane_stern', 'aeroplane_rwing', 'aeroplane_engine',
                      'aeroplane_wheel', 'bicycle_fwheel', 'bicycle_saddle', 'bicycle_handlebar', 'bicycle_chainwheel',
                      'birds_head', 'birds_beak', 'birds_torso', 'birds_neck', 'birds_rwing', 'birds_rleg', 'birds_rfoot',
                      'birds_tail', 'boat', 'bottle_cap', 'bottle_body', 'bus_rightside', 'bus_roofside', 'bus_rightmirror',
                      'bus_fliplate', 'bus_door', 'bus_wheel', 'bus_headlight', 'bus_window', 'car_rightside',
                      'car_roofside', 'car_fliplate', 'car_door', 'car_wheel', 'car_headlight', 'car_window', 'cat_head',
                      'cat_reye', 'cat_rear', 'cat_nose', 'cat_torso', 'cat_neck', 'cat_rfleg', 'cat_rfpa', 'cat_tail',
                      'chair', 'cow_head', 'cow_rear', 'cow_muzzle', 'cow_rhorn', 'cow_torso', 'cow_neck', 'cow_rfuleg',
                      'cow_tail', 'dining_table', 'dog_head', 'dog_reye', 'dog_rear', 'dog_nose', 'dog_torso', 'dog_neck',
                      'dog_rfleg', 'dog_rfpa', 'dog_tail', 'dog_muzzle', 'horse_head', 'horse_rear', 'horse_muzzle',
                      'horse_torso', 'horse_neck', 'horse_rfuleg', 'horse_tail', 'horse_rfho', 'motorbike_fwheel',
                      'motorbike_handlebar', 'motorbike_saddle', 'motorbike_headlight', 'person_head', 'person_reye',
                      'person_rear', 'person_nose', 'person_mouth', 'person_hair', 'person_torso', 'person_neck',
                      'person_ruarm', 'person_rhand', 'person_ruleg', 'person_rfoot', 'pottedplant_pot',
                      'pottedplant_plant', 'sheep_head', 'sheep_rear', 'sheep_muzzle', 'sheep_rhorn', 'sheep_torso',
                      'sheep_neck', 'sheep_rfuleg', 'sheep_tail', 'sofa', 'train_head', 'train_hrightside',
                      'train_hroofside', 'train_headlight', 'train_coach', 'train_crightside', 'train_croofside',
                      'tvmonitor_screen']

    elif num_parts == 58:
        list_parts = ['background', 'aeroplane_body', 'aeroplane_engine', 'aeroplane_wing', 'aeroplane_stern',
                      'aeroplane_wheel', 'bycicle_wheel', 'bycicle_body', 'bird_head', 'bird_wing', 'bird_leg',
                      'bird_torso', 'boat', 'bottle_cap', 'bottle_body', 'bus_window', 'bus_wheel', 'bus_body',
                      'car_window', 'car_wheel', 'car_light', 'car_plate', 'car_body', 'cat_head', 'cat_leg', 'cat_tail',
                      'cat_torso', 'chair', 'cow_head', 'cow_tail', 'cow_leg', 'cow_torso', 'dining_table', 'dog_head',
                      'dog_leg', 'dog_tail', 'dog_torso', 'horse_head', 'horse_tail', 'horse_leg', 'horse_torso',
                      'motorbike_wheel', 'motorbike_body', 'person_head', 'person_torso', 'person_larm', 'person_uarm',
                      'person_lleg', 'person_uleg', 'pottedplant_pot', 'pottedplant_plant', 'sheep_head', 'sheep_leg',
                      'sheep_torso', 'sofa', 'train', 'tvmonitor_screen', 'tvmonitor_frame']

    # Person part
    elif num_parts == 7:
        list_parts = ['background', 'person_head', 'person_torso', 'person_uarm', 'person_larm',
                      'person_uleg', 'person_lleg']

    elif num_parts == 21:
        list_parts = ['background',
                      'airplane',
                      'bicycle',
                      'bird',
                      'boat',
                      'bottle',
                      'bus',
                      'car',
                      'cat',
                      'chair',
                      'cow',
                      'table',
                      'dog',
                      'horse',
                      'motorbike',
                      'person',
                      'potted_plant',
                      'sheep',
                      'sofa',
                      'train',
                      'tv']

    elif num_parts == 2:
        list_parts = ['background',
                      'person']

    else:
        print_message(FilesName.GENERAL_UTILS, "Error: num_parts = " + str(num_parts), MessageType.FAIL)
        exit(-1)

    return list_parts


def list_classes_names(num_classes):
    """
        :return: classe names
    """
    list_parts = []

    if num_classes == 21:
        list_parts = ['background',
                      'airplane',
                      'bicycle',
                      'bird',
                      'boat',
                      'bottle',
                      'bus',
                      'car',
                      'cat',
                      'chair',
                      'cow',
                      'table',
                      'dog',
                      'horse',
                      'motorbike',
                      'person',
                      'potted_plant',
                      'sheep',
                      'sofa',
                      'train',
                      'tv']

    elif num_classes == 2:
        list_parts = ['background',
                      'person']

    else:
        print_message(FilesName.GENERAL_UTILS, "Error: num_classes = " + str(num_classes), MessageType.FAIL)
        exit(-1)

    return list_parts


def create_excel_file(file_name="results", results_classes=None, results_parts=None, path="",
                      num_classes=21, num_parts=108, color_map_path=""):
    """
    :param file_name: name of the file
    :param results_classes: array with IoU for the num_classes classes
    :param results_parts: array with IoU for the # parts
    :param path: where to save the xml file
    :param num_classes:
    :param num_parts: number of parts to predict
    :param color_map_path:
    """

    wb = Workbook()
    dest_filename = path + file_name + '.xlsx'
    ws1 = wb.active
    ws1.title = "results"

    print_message(FilesName.GENERAL_UTILS, "Creating file excel " + dest_filename + " ...")

    pathCMap = color_map_path  # EB: il file cmap dovrà essere diverso in base al numero delle parti prese in esame?
    fileMat = loadmat(pathCMap)
    cmap = fileMat['cmap']

    # color map aRGB hex value
    map = []
    for i in range(len(cmap)):
        value = cmap[i]
        value0 = value[0]
        value1 = value[1]
        value2 = value[2]
        value = ('#{:02x}{:02x}{:02x}'.format(value0, value1, value2))
        map.append(value[1:])

    parts = list_parts_names(num_parts)

    start_index_class_0 = 1
    start_index_class_1 = 1
    start_index_class_2 = 1
    start_index_class_3 = 1
    start_index_class_4 = 1
    start_index_class_5 = 1
    start_index_class_6 = 1
    start_index_class_7 = 1
    start_index_class_8 = 1
    start_index_class_9 = 1
    start_index_class_10 = 1
    start_index_class_11 = 1
    start_index_class_12 = 1
    start_index_class_13 = 1
    start_index_class_14 = 1
    start_index_class_15 = 1
    start_index_class_16 = 1
    start_index_class_17 = 1
    start_index_class_18 = 1
    start_index_class_19 = 1
    start_index_class_20 = 1

    if num_classes == 21:

        if num_parts == 21:
            start_index_class_0 = 1
            start_index_class_1 = 2
            start_index_class_2 = 3
            start_index_class_3 = 4
            start_index_class_4 = 5
            start_index_class_5 = 6
            start_index_class_6 = 7
            start_index_class_7 = 8
            start_index_class_8 = 9
            start_index_class_9 = 10
            start_index_class_10 = 11
            start_index_class_11 = 12
            start_index_class_12 = 13
            start_index_class_13 = 14
            start_index_class_14 = 15
            start_index_class_15 = 16
            start_index_class_16 = 17
            start_index_class_17 = 18
            start_index_class_18 = 19
            start_index_class_19 = 20
            start_index_class_20 = 21

        elif num_parts == 58:
            start_index_class_0 = 1
            start_index_class_1 = 2
            start_index_class_2 = 7
            start_index_class_3 = 9
            start_index_class_4 = 13
            start_index_class_5 = 14
            start_index_class_6 = 16
            start_index_class_7 = 19
            start_index_class_8 = 24
            start_index_class_9 = 28
            start_index_class_10 = 29
            start_index_class_11 = 33
            start_index_class_12 = 34
            start_index_class_13 = 38
            start_index_class_14 = 42
            start_index_class_15 = 44
            start_index_class_16 = 50
            start_index_class_17 = 52
            start_index_class_18 = 55
            start_index_class_19 = 56
            start_index_class_20 = 57

        elif num_parts == 108:
            start_index_class_0 = 1
            start_index_class_1 = 2
            start_index_class_2 = 7
            start_index_class_3 = 11
            start_index_class_4 = 19
            start_index_class_5 = 20
            start_index_class_6 = 22
            start_index_class_7 = 30
            start_index_class_8 = 37
            start_index_class_9 = 46
            start_index_class_10 = 47
            start_index_class_11 = 55
            start_index_class_12 = 56
            start_index_class_13 = 66
            start_index_class_14 = 74
            start_index_class_15 = 78
            start_index_class_16 = 90
            start_index_class_17 = 92
            start_index_class_18 = 100
            start_index_class_19 = 101
            start_index_class_20 = 108

        else:
            print_message(FilesName.GENERAL_UTILS, "Error: num_classes = " + str(num_classes) +
                          " and num_parts = " + str(num_parts), MessageType.FAIL)
            exit(-1)

    elif num_classes == 2:
        # Person parts
        if num_parts == 2:
            start_index_class_0 = 1
            start_index_class_1 = 2

        elif num_parts == 7:
            start_index_class_0 = 1
            start_index_class_1 = 2

        else:
            print_message(FilesName.GENERAL_UTILS, "Error: num_classes = " + str(num_classes) +
                          " and num_parts = " + str(num_parts), MessageType.FAIL)
            exit(-1)

    else:
        print_message(FilesName.GENERAL_UTILS, "Error: num_classes = " + str(num_classes) +
                      " and num_parts = " + str(num_parts), MessageType.FAIL)
        exit(-1)

    map_part = []

    if num_classes == 2:

        map_part.append(start_index_class_0)
        map_part.append(start_index_class_1)

        ws1.merge_cells(start_row=start_index_class_0, end_row=start_index_class_1 - 1, end_column=1, start_column=1)
        ws1.merge_cells(start_row=start_index_class_1, end_row=num_parts, end_column=1, start_column=1)

        ws1.merge_cells(start_row=start_index_class_0, end_row=start_index_class_1 - 1, end_column=2, start_column=2)
        ws1.merge_cells(start_row=start_index_class_1,end_row=num_parts, end_column=2, start_column=2)

    elif num_classes == 21:

        map_part.append(start_index_class_0)
        map_part.append(start_index_class_1)
        map_part.append(start_index_class_2)
        map_part.append(start_index_class_3)
        map_part.append(start_index_class_4)
        map_part.append(start_index_class_5)
        map_part.append(start_index_class_6)
        map_part.append(start_index_class_7)
        map_part.append(start_index_class_8)
        map_part.append(start_index_class_9)
        map_part.append(start_index_class_10)
        map_part.append(start_index_class_11)
        map_part.append(start_index_class_12)
        map_part.append(start_index_class_13)
        map_part.append(start_index_class_14)
        map_part.append(start_index_class_15)
        map_part.append(start_index_class_16)
        map_part.append(start_index_class_17)
        map_part.append(start_index_class_18)
        map_part.append(start_index_class_19)
        map_part.append(start_index_class_20)

        ws1.merge_cells(start_row=start_index_class_0, end_row=start_index_class_1 - 1, end_column=1, start_column=1)
        ws1.merge_cells(start_row=start_index_class_1, end_row=start_index_class_2 - 1, end_column=1, start_column=1)
        ws1.merge_cells(start_row=start_index_class_2, end_row=start_index_class_3 - 1, end_column=1, start_column=1)
        ws1.merge_cells(start_row=start_index_class_3, end_row=start_index_class_4 - 1, end_column=1, start_column=1)
        ws1.merge_cells(start_row=start_index_class_4, end_row=start_index_class_5 - 1, end_column=1, start_column=1)
        ws1.merge_cells(start_row=start_index_class_5, end_row=start_index_class_6 - 1, end_column=1, start_column=1)
        ws1.merge_cells(start_row=start_index_class_6, end_row=start_index_class_7 - 1, end_column=1, start_column=1)
        ws1.merge_cells(start_row=start_index_class_7, end_row=start_index_class_8 - 1, end_column=1, start_column=1)
        ws1.merge_cells(start_row=start_index_class_8, end_row=start_index_class_9 - 1, end_column=1, start_column=1)
        ws1.merge_cells(start_row=start_index_class_9, end_row=start_index_class_10 - 1, end_column=1, start_column=1)
        ws1.merge_cells(start_row=start_index_class_10, end_row=start_index_class_11 - 1, end_column=1, start_column=1)
        ws1.merge_cells(start_row=start_index_class_11, end_row=start_index_class_12 - 1, end_column=1, start_column=1)
        ws1.merge_cells(start_row=start_index_class_12, end_row=start_index_class_13 - 1, end_column=1, start_column=1)
        ws1.merge_cells(start_row=start_index_class_13, end_row=start_index_class_14 - 1, end_column=1, start_column=1)
        ws1.merge_cells(start_row=start_index_class_14, end_row=start_index_class_15 - 1, end_column=1, start_column=1)
        ws1.merge_cells(start_row=start_index_class_15, end_row=start_index_class_16 - 1, end_column=1, start_column=1)
        ws1.merge_cells(start_row=start_index_class_16, end_row=start_index_class_17 - 1, end_column=1, start_column=1)
        ws1.merge_cells(start_row=start_index_class_17, end_row=start_index_class_18 - 1, end_column=1, start_column=1)
        ws1.merge_cells(start_row=start_index_class_18, end_row=start_index_class_19 - 1, end_column=1, start_column=1)
        ws1.merge_cells(start_row=start_index_class_19, end_row=start_index_class_20 - 1, end_column=1, start_column=1)
        ws1.merge_cells(start_row=start_index_class_20, end_row=num_parts, end_column=1, start_column=1)

        ws1.merge_cells(start_row=start_index_class_0, end_row=start_index_class_1 - 1, end_column=2, start_column=2)
        ws1.merge_cells(start_row=start_index_class_1, end_row=start_index_class_2 - 1, end_column=2, start_column=2)
        ws1.merge_cells(start_row=start_index_class_2, end_row=start_index_class_3 - 1, end_column=2, start_column=2)
        ws1.merge_cells(start_row=start_index_class_3, end_row=start_index_class_4 - 1, end_column=2, start_column=2)
        ws1.merge_cells(start_row=start_index_class_4, end_row=start_index_class_5 - 1, end_column=2, start_column=2)
        ws1.merge_cells(start_row=start_index_class_5, end_row=start_index_class_6 - 1, end_column=2, start_column=2)
        ws1.merge_cells(start_row=start_index_class_6, end_row=start_index_class_7 - 1, end_column=2, start_column=2)
        ws1.merge_cells(start_row=start_index_class_7, end_row=start_index_class_8 - 1, end_column=2, start_column=2)
        ws1.merge_cells(start_row=start_index_class_8, end_row=start_index_class_9 - 1, end_column=2, start_column=2)
        ws1.merge_cells(start_row=start_index_class_9, end_row=start_index_class_10 - 1, end_column=2, start_column=2)
        ws1.merge_cells(start_row=start_index_class_10, end_row=start_index_class_11 - 1, end_column=2, start_column=2)
        ws1.merge_cells(start_row=start_index_class_11, end_row=start_index_class_12 - 1, end_column=2, start_column=2)
        ws1.merge_cells(start_row=start_index_class_12, end_row=start_index_class_13 - 1, end_column=2, start_column=2)
        ws1.merge_cells(start_row=start_index_class_13, end_row=start_index_class_14 - 1, end_column=2, start_column=2)
        ws1.merge_cells(start_row=start_index_class_14, end_row=start_index_class_15 - 1, end_column=2, start_column=2)
        ws1.merge_cells(start_row=start_index_class_15, end_row=start_index_class_16 - 1, end_column=2, start_column=2)
        ws1.merge_cells(start_row=start_index_class_16, end_row=start_index_class_17 - 1, end_column=2, start_column=2)
        ws1.merge_cells(start_row=start_index_class_17, end_row=start_index_class_18 - 1, end_column=2, start_column=2)
        ws1.merge_cells(start_row=start_index_class_18, end_row=start_index_class_19 - 1, end_column=2, start_column=2)
        ws1.merge_cells(start_row=start_index_class_19, end_row=start_index_class_20 - 1, end_column=2, start_column=2)
        ws1.merge_cells(start_row=start_index_class_20, end_row=num_parts, end_column=2, start_column=2)

    else:
        print_message(FilesName.GENERAL_UTILS, "Error: num_classes = " + str(num_classes) +
                      " and num_parts = " + str(num_parts), MessageType.FAIL)
        exit(-1)

    classes = list_classes_names(num_classes=num_classes)
    index_class = 0
    for row in map_part:
        cell = ws1.cell(column=1, row=row, value="{0}".format(classes[index_class]))

        if results_classes is not None:
            _ = ws1.cell(column=2, row=row, value="{0}".format(results_classes[index_class]))

        if index_class != 0:
            cell.fill = PatternFill("solid", fgColor=(map[index_class]))

        index_class = index_class + 1

    for row in range(len(parts)):
        cell = ws1.cell(column=3, row=row + 1, value="{0}".format(parts[row]))
        if results_parts is not None:
            _ = ws1.cell(column=4, row=row + 1, value="{0}".format(results_parts[row]))

    print_message(FilesName.GENERAL_UTILS, "Saving excel file...")

    wb.save(filename=dest_filename)


def compute_and_print_IoU_per_class(confusion_matrix, num_classes, num_parts, part_mask=None):
    """
    Computes and prints mean intersection over union divided per class
    :param confusion_matrix: confusion matrix needed for the computation
    :param num_classes:
    :param num_parts:
    :param part_mask:
    :param name_part:
    """

    name_part = list_parts_names(num_parts)

    print_message(FilesName.GENERAL_UTILS, "compute_and_print_IoU_per_class", MessageType.HEADER)

    map_parts = map_parts_to_classes(num_classes, num_parts)

    classes_names = list_classes_names(num_classes)

    mIoU = 0
    mIoU_nobackgroud = 0
    IoU_per_part = np.zeros([num_parts], np.float32)
    true_parts = 0

    per_part_pixel_acc = np.zeros([num_parts], np.float32)

    mean_part_acc_num = 0

    true_parts_pix = 0
    mean_part_acc_den = 0

    mean_part_acc_num_nobgr = 0
    mean_part_acc_den_nobgr = 0
    mean_part_acc_sum_nobgr = 0
    mean_part_acc_sum = 0

    if part_mask is None:
        part_mask = np.ones([num_parts], np.int8)

    for part_index in range(num_parts):

        if part_mask[part_index] == 1:
            # IoU = true_positive / (true_positive + false_positive + false_negative)
            TP = confusion_matrix[part_index, part_index]
            FP = np.sum(confusion_matrix[:, part_index]) - TP
            FN = np.sum(confusion_matrix[part_index]) - TP
            # TN = np.sum(confusion_matrix) - TP - FP - FN

            denominator = (TP + FP + FN)
            # If the denominator is 0, we need to ignore the class.
            if denominator == 0:
                denominator = 1
                print_message(FilesName.GENERAL_UTILS,
                              "Ignore part " + name_part[part_index] + "... IoU denominator is 0",
                              MessageType.WARNING)
            else:
                true_parts += 1

            # per-part pixel accuracy
            if not TP == 0:
                # if not np.isnan(TP):
                tmp = (TP + FN)
                per_part_pixel_acc[part_index] = TP / tmp

            IoU = TP / denominator
            IoU_per_part[part_index] += IoU
            mIoU += IoU

            if part_index > 0:
                mIoU_nobackgroud += IoU

            # mean class accuracy
            if not np.isnan(per_part_pixel_acc[part_index]):
                mean_part_acc_num += TP
                mean_part_acc_den += TP + FN

                mean_part_acc_sum += per_part_pixel_acc[part_index]
                true_parts_pix += 1

                if part_index > 0:
                    mean_part_acc_num_nobgr += TP
                    mean_part_acc_den_nobgr += TP + FN
                    mean_part_acc_sum_nobgr += per_part_pixel_acc[part_index]

    mIoU = mIoU / true_parts
    mIoU_nobackgroud = mIoU_nobackgroud / (true_parts - 1)

    mean_part_pix_acc = mean_part_acc_num / mean_part_acc_den
    mean_pixel_acc_nobackground = mean_part_acc_num_nobgr / mean_part_acc_den_nobgr

    # Stats per part
    print_message(FilesName.GENERAL_UTILS, "---------------------------------------------------------------------------")
    print_message(FilesName.GENERAL_UTILS, "True_parts: " + str(true_parts))
    print_message(FilesName.GENERAL_UTILS, "---------------------------------------------------------------------------")
    print_message(FilesName.GENERAL_UTILS, "-- background --")
    print_message(FilesName.GENERAL_UTILS, "IoU for part background: " + str(IoU_per_part[0] * 100))
    print_message(FilesName.GENERAL_UTILS, "Pixel Accuracy for part background: " + str(per_part_pixel_acc[0] * 100))
    print_message(FilesName.GENERAL_UTILS, "---------------------------------------------------------------------------")

    parts_not_found = []

    for part in range(1, num_parts):
        if IoU_per_part[part] > 0:
            print_message(FilesName.GENERAL_UTILS, "-- " + str(part) + " -- " + name_part[part] + " --")
            print_message(FilesName.GENERAL_UTILS,
                          "IoU for part " + name_part[part] + ": " + str(IoU_per_part[part] * 100))
            print_message(FilesName.GENERAL_UTILS,
                          "Pixel Accuracy for part " + str(part) + ": " + str(per_part_pixel_acc[part] * 100))
            print_message(FilesName.GENERAL_UTILS,
                          "---------------------------------------------------------------------------")
        else:
            parts_not_found.append(part)

    # Stats medie per le parti
    print_message(FilesName.GENERAL_UTILS, " ")
    print_message(FilesName.GENERAL_UTILS, "--METRICS--")
    print_message(FilesName.GENERAL_UTILS,
                  "Mean true parts accuracy: " + str((mean_part_acc_sum / true_parts_pix) * 100))
    print_message(FilesName.GENERAL_UTILS, "Mean pixel accuracy: " + str(mean_part_pix_acc * 100))
    print_message(FilesName.GENERAL_UTILS,
                  "Mean pixel accuracy no background : " + str(mean_pixel_acc_nobackground * 100))
    print_message(FilesName.GENERAL_UTILS, "mIoU: " + str(mIoU * 100))
    print_message(FilesName.GENERAL_UTILS, "mIoU no backgroud: " + str(mIoU_nobackgroud * 100))
    print_message(FilesName.GENERAL_UTILS, "---------------------------------------------------------------------------")

    for part in range(len(parts_not_found)):
        print_message(FilesName.GENERAL_UTILS,
                      "Part not found " + str(parts_not_found[part]) + "_" + name_part[parts_not_found[part]],
                      MessageType.WARNING)

    print_message(FilesName.GENERAL_UTILS,
                  "Parts_" + str(num_parts - len(parts_not_found)) + "/" + str(num_parts) + "\n")

    # Stats per class
    per_class_m_iou = np.zeros([num_classes], np.float32)
    parts_per_class = np.zeros([num_classes], np.int32)
    per_class_pixel_acc = np.zeros([num_classes], np.float32)

    for class_index in range(0, num_classes):
        parts = map_parts[class_index]

        for part_index in range(parts[0], parts[1]):
            per_class_m_iou[class_index] += IoU_per_part[part_index]
            per_class_pixel_acc[class_index] += per_part_pixel_acc[part_index]
            parts_per_class[class_index] += 1

        per_class_m_iou[class_index] = per_class_m_iou[class_index] / parts_per_class[class_index]
        per_class_pixel_acc[class_index] = per_class_pixel_acc[class_index] / parts_per_class[class_index]

        print_message(FilesName.GENERAL_UTILS, "-- " + str(class_index) + " -- " + classes_names[class_index] + " --")
        print_message(FilesName.GENERAL_UTILS,
                      "IoU for class " + classes_names[class_index] + ": " + str(per_class_m_iou[class_index] * 100))
        print_message(FilesName.GENERAL_UTILS,
                      "Pixel Accuracy for class " + str(class_index) + ": " + str(per_class_pixel_acc[class_index] * 100))
        print_message(FilesName.GENERAL_UTILS,
                      "---------------------------------------------------------------------------")

    # Stats medie per le classi
    mean_class_accuracy = 0
    mean_class_iou = 0
    for class_index in range(0, num_classes):
        mean_class_accuracy += per_class_pixel_acc[class_index]
        mean_class_iou += per_class_m_iou[class_index]

    mean_class_accuracy = mean_class_accuracy / num_classes
    mean_class_iou = mean_class_iou / num_classes

    print_message(FilesName.GENERAL_UTILS, " ")
    print_message(FilesName.GENERAL_UTILS, "--METRICS--")
    print_message(FilesName.GENERAL_UTILS, "Mean pixel accuracy for all classes: " + str(mean_class_accuracy * 100))
    print_message(FilesName.GENERAL_UTILS, "mIoU for all classes: " + str(mean_class_iou * 100))
    print_message(FilesName.GENERAL_UTILS, "---------------------------------------------------------------------------")

    # return
    return mIoU * 100, IoU_per_part
